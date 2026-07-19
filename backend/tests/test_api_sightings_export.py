"""
Tests for the Vogelwarte (RING) Wiederfunde Excel export endpoint.
"""

import io
from datetime import date
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from src.database.models import Sighting
from src.database.user_models import User


EXPORT_URL = "/api/sightings/export/vogelwarte"


@pytest.fixture
def dev_org_id(client, test_db):
    """Warm up the dev user (created on first authed request) and return its org_id.

    In dev mode the endpoint scopes to this org, so test sightings must share it.
    """
    client.get("/api/sightings/count")
    user = test_db.query(User).filter(User.email == "dev@vogelring.local").first()
    return user.org_id

EXPECTED_HEADERS = [
    "Datum",
    "Ort",
    "RING-Ort",
    "Lat",
    "Lon",
    "Ring",
    "Spezies",
    "Alter",
    "Geschlecht",
    "Status",
    "Bemerkungen",
]

# 0-indexed positions of the columns the tests assert on.
COL_RING = 5
COL_AGE = 7
COL_STATUS = 9
COL_BEMERKUNGEN = 10


def _add(test_db, org_id, **kwargs):
    sighting = Sighting(id=uuid4(), org_id=org_id, **kwargs)
    test_db.add(sighting)
    test_db.commit()
    return sighting


def _load_rows(response):
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


class TestVogelwarteExport:
    def test_empty_export_has_only_header(self, client):
        response = client.get(EXPORT_URL)
        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]

        rows = _load_rows(response)
        assert rows == [EXPECTED_HEADERS]

    def test_filtering_and_transforms(self, client, test_db, dev_org_id):
        # Included: MG, empty age/sex -> defaults
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 3, 1),
            melded=False,
            status="MG",
            age=None,
            sex=None,
            place="Kalbach",
            ring="R1",
            species="Larus ridibundus",
            melder="Obs A",
            comment="c1",
            lat=50.1,
            lon=8.6,
        )
        # Included: BV, populated age/sex pass through; melded NULL treated as not reported
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 2, 1),
            melded=None,
            status="BV",
            age=6,  # RING code (Adult)
            sex=1,  # männlich
            place="Nied",
            ring="R2",
            species="Anas platyrhynchos",
            melder="Obs B",
            comment="c2",
        )
        # Included: other status -> "unbekannt / nicht erfasst"
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 5, 1),
            melded=False,
            status="NB",
            place="Mainkai",
            ring="R3",
        )
        # Excluded: already reported
        _add(test_db, dev_org_id, date=date(2026, 4, 1), melded=True, status="MG", ring="X1")
        # Excluded: before start date
        _add(test_db, dev_org_id, date=date(2025, 12, 31), melded=False, status="MG", ring="X2")

        response = client.get(EXPORT_URL)
        assert response.status_code == 200
        rows = _load_rows(response)

        assert rows[0] == EXPECTED_HEADERS
        data = rows[1:]
        # 3 included, ordered by date ascending
        assert len(data) == 3

        # Row 1: 2026-02-01, BV. Place "Nied" is not in the RING lookup, so
        # RING-Ort/Lat/Lon are blank (read back as None). Melder + comment fold
        # into Bemerkungen (Melder kept because it's not "IR").
        assert data[0] == [
            "01.02.2026",
            "Nied",
            None,  # RING-Ort
            None,  # Lat
            None,  # Lon
            "R2",
            "Anas platyrhynchos",
            "6 Älter als vorjährig",  # code 6 -> label
            "männlich",
            "nestbauend oder brütend",
            "Melder: Obs B / Kommentare: c2",  # Bemerkungen
        ]
        # Row 2: 2026-03-01, MG, empty age/sex defaults
        assert data[1] == [
            "01.03.2026",
            "Kalbach",
            None,  # RING-Ort
            None,  # Lat
            None,  # Lon
            "R1",
            "Larus ridibundus",
            "2 Fängling",
            "unbekannt",
            "in Mausertrupp",
            "Melder: Obs A / Kommentare: c1",  # Bemerkungen
        ]
        # Row 3: 2026-05-01, NB -> unknown status
        assert data[2][0] == "01.05.2026"
        assert data[2][COL_STATUS] == "unbekannt / nicht erfasst"

        # The sighting's OWN stored coordinates must NOT leak — only RING-lookup
        # coordinates are exported, and these test places aren't in the lookup.
        flat = [str(c) for r in rows for c in r if c is not None]
        assert not any("50.1" in v or "8.6" in v for v in flat)

    def test_start_date_override(self, client, test_db, dev_org_id):
        _add(test_db, dev_org_id, date=date(2026, 1, 15), melded=False, status="MG", ring="A")
        _add(test_db, dev_org_id, date=date(2026, 6, 15), melded=False, status="MG", ring="B")

        response = client.get(EXPORT_URL, params={"start_date": "2026-06-01"})
        assert response.status_code == 200
        rows = _load_rows(response)
        assert len(rows) == 2  # header + 1 row
        assert rows[1][COL_RING] == "B"

    def test_end_date_bound(self, client, test_db, dev_org_id):
        _add(test_db, dev_org_id, date=date(2026, 1, 15), melded=False, status="MG", ring="A")
        _add(test_db, dev_org_id, date=date(2026, 6, 15), melded=False, status="MG", ring="B")

        response = client.get(
            EXPORT_URL, params={"start_date": "2026-01-01", "end_date": "2026-03-01"}
        )
        assert response.status_code == 200
        rows = _load_rows(response)
        assert len(rows) == 2  # header + 1 row (only the January one)
        assert rows[1][COL_RING] == "A"

    def test_age_labels(self, client, test_db, dev_org_id):
        # Post-migration the DB holds integer codes; the export labels them.
        cases = [
            (None, "2 Fängling"),  # no age -> Fängling for RING
            (1, "1 Nestling"),
            (2, "2 Fängling"),
            (3, "3 Diesjährig"),
            (5, "5 Vorjährig"),
            (6, "6 Älter als vorjährig"),
        ]
        for i, (code, _exp) in enumerate(cases):
            _add(
                test_db,
                dev_org_id,
                date=date(2026, 1, 1),
                melded=False,
                status="MG",
                age=code,
                ring=f"AGE{i}",
                place=f"p{i:02d}",  # order rows deterministically by place
            )
        rows = _load_rows(client.get(EXPORT_URL))
        got = {r[COL_RING]: r[COL_AGE] for r in rows[1:]}  # ring -> age column
        for i, (_code, exp) in enumerate(cases):
            assert got[f"AGE{i}"] == exp

    def test_place_matched_to_ring_place_and_coords(self, client, test_db, dev_org_id):
        # "F, Mainkai, Innenstadt" is a mapped Vogelring place in the RING lookup.
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 3, 1),
            melded=False,
            status="MG",
            ring="M1",
            place="F, Mainkai, Innenstadt",
        )
        rows = _load_rows(client.get(EXPORT_URL))
        row = rows[1]
        assert row[2].startswith("Frankfurt, Mainkai")  # RING-Ort
        assert row[3] == 50.11194  # Lat from the lookup
        assert row[4] == 8.7025  # Lon from the lookup

    def test_unmapped_place_smart_matched_and_flagged(self, client, test_db, dev_org_id):
        # "F, Bethmannweiher" is NOT in Ingo's explicit map, but the sighting's
        # coordinates sit ~3 m from RING "Bethmann Weiher" -> auto-suggested.
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 3, 1),
            melded=False,
            status="MG",
            ring="S1",
            place="F, Bethmannweiher",
            lat=50.11775,
            lon=8.69082,
        )
        row = _load_rows(client.get(EXPORT_URL))[1]
        assert "Bethmann" in row[2]
        assert row[2].endswith("(auto)")  # flagged as unverified suggestion
        assert isinstance(row[3], float) and isinstance(row[4], float)

    def test_unmapped_place_without_close_match_stays_blank(self, client, test_db, dev_org_id):
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 3, 1),
            melded=False,
            status="MG",
            ring="S2",
            place="F, Erlenbruch",
            lat=50.12912,
            lon=8.72790,
        )
        row = _load_rows(client.get(EXPORT_URL))[1]
        assert row[2] is None  # RING-Ort blank
        assert row[3] is None and row[4] is None

    def test_bemerkungen_bundles_non_ring_fields(self, client, test_db, dev_org_id):
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 3, 1),
            melded=False,
            status="MG",
            ring="B1",
            place="Nirgendwo",
            pair="F",  # Familie
            partner="281937",
            breed_size=8,
            family_size=2,
        )
        rows = _load_rows(client.get(EXPORT_URL))
        bemerkungen = rows[1][COL_BEMERKUNGEN]
        assert bemerkungen == (
            "Familien Status: Familie / Partner: 281937 / "
            "Nicht flügge Junge: 8 / Flügge Junge: 2"
        )

    def test_bemerkungen_order_and_melder_ir_dropped(self, client, test_db, dev_org_id):
        # Melder "IR" is dropped; all other fields appear in Ingo's specified order.
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 3, 1),
            melded=False,
            status="MG",
            ring="B2",
            place="Nirgendwo",
            melder="IR",  # must NOT appear
            large_group_size=85,
            small_group_size=19,
            pair="F",
            partner="281937",
            breed_size=8,
            family_size=2,
            field_fruit="Rasen",
            comment="ruht am Ufer",
        )
        b = _load_rows(client.get(EXPORT_URL))[1][COL_BEMERKUNGEN]
        assert b == (
            "Großgruppe: 85 / Kleingruppe: 19 / Familien Status: Familie / "
            "Partner: 281937 / Nicht flügge Junge: 8 / Flügge Junge: 2 / "
            "Feldfrucht: Rasen / Kommentare: ruht am Ufer"
        )

    def test_bemerkungen_keeps_non_ir_melder(self, client, test_db, dev_org_id):
        _add(
            test_db,
            dev_org_id,
            date=date(2026, 3, 1),
            melded=False,
            status="MG",
            ring="B3",
            place="Nirgendwo",
            melder="OR",
        )
        b = _load_rows(client.get(EXPORT_URL))[1][COL_BEMERKUNGEN]
        assert b == "Melder: OR"

    def test_export_does_not_change_melded(self, client, test_db, dev_org_id):
        s = _add(
            test_db, dev_org_id, date=date(2026, 3, 1), melded=False, status="MG", ring="R1"
        )
        client.get(EXPORT_URL)
        test_db.refresh(s)
        assert s.melded is False
